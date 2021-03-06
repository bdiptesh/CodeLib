"""
utils.

Utlities module
-----------------

**Available functions:**

- :``elapsed_time``: Function to return elapsed time.
- :``prime_list``: Generate a list of prime numbers till limit.
- :``flat_list``: Flatten a deeply nested list.
- :``append_df_to_excel``: Append pandas.DataFrame to existing excel workbook.
- :``df_size``: Determine size of a pandas.DataFrame.
- :``table_output``: Function to return prettytable object.

Author
------
::

    Author: Diptesh Basak
    Date: Apr 06, 2020
    License: BSD 3-Clause
"""

# =============================================================================
# --- Import libraries
# =============================================================================

import time
import copy

from typing import List, Any

import pandas as pd
import numpy as np

from prettytable import PrettyTable
from openpyxl import load_workbook

# =============================================================================
# --- DO NOT CHANGE ANYTHING FROM HERE
# =============================================================================

# pylint: disable=invalid-name
# pylint: disable=abstract-class-instantiated

# =============================================================================
# --- User defined functions
# =============================================================================


def elapsed_time(text: str,
                 start_t: int,
                 sept: int = 70
                 ) -> str:
    """
    Return elapsed time.

    Parameters
    ----------
    :text: str

        Text to be printed

    :start_t: int

        Generated from time.time_ns()

    :sept: int

        Length of text

    Returns
    -------
    str
        A string containing arg "text" followed by hours, minutes, seconds,
        milliseconds.

    Example usage
    -------------

    >>> import time
    >>> start = time.time_ns()
    >>> time.sleep(2)
    >>> elapsed_time("Time taken:", start)
    Time taken: 00:00:02 000 ms

    """
    second, ms = divmod(round((time.time_ns() / 1e6) - (start_t / 1e6), 0),
                        1000)
    minute, second = divmod(second, 60)
    hour, minute = divmod(minute, 60)
    fn_op = text + str("%02d:%02d:%02d %03d ms" % (hour, minute, second, ms))\
        .rjust(sept - len(text))
    return fn_op


def prime_list(limit: int = 2) -> List[int]:
    """
    Generate a list of prime numbers till limit.

    Parameters
    ----------
    :limit: int

        An integer till which we need to generate primes

    Returns
    -------
    list
        A list containing prime numbers from 2 to limit

    Example usage
    -------------
    >>> prime_numbers = prime_list(20)

    """
    is_prime = np.ones(limit + 1, dtype=np.bool)
    for n in range(2, int(limit**0.5 + 1.5)):
        if is_prime[n]:
            is_prime[n*n::n] = 0
    return list(np.nonzero(is_prime)[0][2:])


def append_df_to_excel(filename: str,
                       df: pd.DataFrame,
                       sheet_name: str = 'Sheet1',
                       startrow: int = None,
                       **to_excel_kwargs: Any
                       ) -> None:
    """
    Append dataframe to spreadsheet.

    Append a DataFrame **df** to existing Excel file **filename**
    into **sheet_name** Sheet.
    If **filename** doesn't exist, then this function will create it.

    Parameters
    ----------
    :filename: str

        File path or existing ExcelWriter. Example: '/path/to/file.xlsx'

    :df: pandas.DataFrame

        dataframe to save to workbook

    :sheet_name: str, optional, default : 'Sheet1'

        Name of sheet which will contain DataFrame

    :startrow: int, optional, default : None

        upper left cell row to dump data frame.
        calculate the last row in the existing DF and write to the next row.

    :to_excel_kwargs:

        arguments which will be passed to `DataFrame.to_excel`. It
        can be dictionary.

    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def df_size(df: pd.DataFrame) -> str:
    """
    Determine size of a pandas.DataFrame.

    Parameters
    ----------
    :df: pandas.DataFrame

    """
    memorySize = df.memory_usage(index=True).sum()
    if memorySize >= 1024.0 ** 4:
        op = str(np.around(memorySize / (1024.0 ** 4), decimals=2)) + " TB"
    elif memorySize >= 1024.0 ** 3:
        op = str(np.around(memorySize / (1024.0 ** 3), decimals=2)) + " GB"
    elif memorySize >= 1024.0 ** 2:
        op = str(np.around(memorySize / (1024.0 ** 2), decimals=2)) + " MB"
    elif memorySize >= 1024.0:
        op = str(np.around(memorySize / (1024.0), decimals=2)) + " KB"
    else:
        op = str(np.around(memorySize, decimals=2)) + " Bytes"
    return op


def table_output(df: pd.DataFrame,
                 col: List[str],
                 header: List[str],
                 precision: int = 3
                 ) -> PrettyTable():
    """
    Return prettytable object.

    Parameters
    ----------
    :df: pandas.DataFrame

        Input pandas dataframe

    :col: list

        A list containing the columns to be returned

    :header: list

        A list containing the headers required in table output

    :precision: int, `optional`, `default:` ``3``

        Decimal points to be printed in table output

    """
    pt_ip = copy.copy(df)
    pt_ip = pt_ip[col]
    pt_ip = pt_ip.round(decimals=precision)
    pt_op = PrettyTable()
    pt_op.field_names = header
    for j in enumerate(header):
        if j[0] == 0:
            pt_op.align[j[1]] = "l"
        else:
            pt_op.align[j[1]] = "r"
    for i in pt_ip.index:
        pt_op.add_row(pt_ip.loc[i, :])
    return pt_op
